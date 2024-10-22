# script3_test_applications.py

import os
import sys
import time
import psutil
import tkinter as tk
from tkinter import filedialog, ttk
from tkinter.scrolledtext import ScrolledText
from pywinauto import Desktop
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import logging
import re

# Configure logging to file only for initial setup
logging.basicConfig(level=logging.INFO, filename='application_test.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Global variables for the progress window
progress_window = None
progress_label = None
progress_bar = None
log_text_widget = None

def load_files(shortcuts_file_path, executables_file_path):
    if not os.path.exists(shortcuts_file_path):
        logging.error(f"Shortcuts file not found at {shortcuts_file_path}.")
        return None, None

    if not os.path.exists(executables_file_path):
        logging.error(f"Executables file not found at {executables_file_path}.")
        return None, None

    with open(shortcuts_file_path, 'r', encoding='utf-8') as f:
        shortcuts = [line.strip() for line in f]

    with open(executables_file_path, 'r', encoding='utf-8') as f:
        executables = [line.strip() for line in f]

    # Filter out folder entries and ensure lengths match
    filtered_shortcuts = []
    filtered_executables = []

    for shortcut, executable in zip(shortcuts, executables):
        if not shortcut.startswith("[Folder]"):
            filtered_shortcuts.append(shortcut)
            filtered_executables.append(executable)

    if len(filtered_shortcuts) != len(filtered_executables):
        logging.error("The number of shortcuts and executables does not match after filtering.")
        return None, None

    logging.info("Loaded shortcuts and executables files successfully.")
    return filtered_shortcuts, filtered_executables

def handle_uac_prompt():
    try:
        # UAC prompts run on a separate desktop and are difficult to interact with
        # But we can detect if the UAC process is running
        for proc in psutil.process_iter(['pid', 'name']):
            if proc.info['name'] == 'Consent.exe':
                logging.warning("Detected UAC prompt (Consent.exe is running).")
                return True
        return False
    except Exception as e:
        logging.error(f"Error detecting UAC prompt: {e}")
        return False

def kill_process_tree(pid, including_parent=True):
    terminated_executables = []
    try:
        parent = psutil.Process(pid)
        parent_name = parent.name()
        if parent_name.lower() in ['svchost.exe', 'csrss.exe', 'wininit.exe', 'services.exe']:
            logging.warning(f"Skipping termination of system process: PID {parent.pid}, Name {parent_name}")
            return terminated_executables
        children = parent.children(recursive=True)
        for child in children:
            try:
                child_name = child.name()
                if child_name.lower() in ['svchost.exe', 'csrss.exe', 'wininit.exe', 'services.exe']:
                    logging.warning(f"Skipping termination of system process: PID {child.pid}, Name {child_name}")
                    continue
                logging.info(f"Terminating child process: PID {child.pid}, Name {child_name}")
                child.terminate()
                terminated_executables.append(child_name)
            except psutil.NoSuchProcess:
                pass
        gone, still_alive = psutil.wait_procs(children, timeout=5)
        for child in still_alive:
            try:
                child_name = child.name()
                if child_name.lower() in ['svchost.exe', 'csrss.exe', 'wininit.exe', 'services.exe']:
                    logging.warning(f"Skipping killing of system process: PID {child.pid}, Name {child_name}")
                    continue
                logging.info(f"Killing child process: PID {child.pid}, Name {child_name}")
                child.kill()
                terminated_executables.append(child_name)
            except psutil.NoSuchProcess:
                pass
        if including_parent:
            try:
                logging.info(f"Terminating parent process: PID {parent.pid}, Name {parent_name}")
                parent.terminate()
                parent.wait(5)
                terminated_executables.append(parent_name)
            except psutil.NoSuchProcess:
                pass
    except psutil.NoSuchProcess:
        pass
    return terminated_executables

def create_progress_window(total_apps):
    global progress_window, progress_label, progress_bar, log_text_widget
    progress_window = tk.Toplevel()
    progress_window.title("Application Testing Progress")
    progress_window.geometry("400x300+1200+50")  # Adjust position and size as needed
    progress_window.attributes('-topmost', True)
    progress_window.resizable(False, False)

    progress_label = tk.Label(progress_window, text="Starting tests...", font=("Helvetica", 12))
    progress_label.pack(padx=10, pady=5)

    progress_bar = ttk.Progressbar(progress_window, length=380, mode='determinate', maximum=total_apps)
    progress_bar.pack(padx=10, pady=5)

    # Add ScrolledText widget for log display
    log_text_widget = ScrolledText(progress_window, width=48, height=12, state='disabled', font=("Courier", 8))
    log_text_widget.pack(padx=10, pady=5)

def update_progress_window(current_app_index, app_name):
    progress_label.config(text=f"Testing application {current_app_index}/{int(progress_bar['maximum'])}: {app_name}")
    progress_bar['value'] = current_app_index
    progress_window.update_idletasks()

def log_to_text_widget(message):
    if log_text_widget is not None:
        log_text_widget.configure(state='normal')
        log_text_widget.insert(tk.END, message + '\n')
        log_text_widget.see(tk.END)
        log_text_widget.configure(state='disabled')
        progress_window.update_idletasks()

# Custom logging handler
class TextWidgetHandler(logging.Handler):
    def emit(self, record):
        msg = self.format(record)
        log_to_text_widget(msg)

def launch_and_test_application(shortcut_path, expected_exe_path, app_name):
    # Extract file names for shortcut and expected executable
    shortcut_name = os.path.basename(shortcut_path)
    expected_executable_name = os.path.basename(expected_exe_path).lower()

    result = {
        'Name': app_name,
        'Shortcut Path': shortcut_name,
        'Expected Executable': expected_executable_name,
        'Status': 'Not Tested',
        'Remarks': '',
        'Associated Windows': '',
        'Terminated Executables': '',
        'Closed Windows': ''
    }

    # Mapping of expected executable names to actual executable names
    executable_aliases = {
        'cmd.exe': 'WindowsTerminal.exe',
        'powershell.exe': 'WindowsTerminal.exe',
        'wmplayer.exe': 'setup_wm.exe',
        # Add other mappings if necessary
    }

    actual_executable_name = executable_aliases.get(expected_executable_name, expected_executable_name)

    associated_window_titles = []
    terminated_executables = []  # List to store terminated executables
    closed_windows = []  # List to store closed window titles

    logging.info(f"Starting test for application: {app_name}")

    if not os.path.exists(shortcut_path):
        result['Status'] = 'Failed'
        result['Remarks'] = f'Shortcut not found: {shortcut_name}'
        logging.error(f"Shortcut not found: {shortcut_name}")
        return result

    try:
        # Record initial processes and windows
        processes_before = {p.pid for p in psutil.process_iter(['pid'])}
        windows_before = set(w.handle for w in Desktop(backend="uia").windows())
        logging.info("Recorded initial processes and windows.")

        # Start the application using os.startfile()
        os.startfile(shortcut_path)
        logging.info(f"Launched application using shortcut: {shortcut_name}")

        # Wait for the application to initialize or until a new window appears
        max_wait_time = 20  # Increased to accommodate slow-loading applications
        poll_interval = 2   # Time between polls (in seconds)
        start_time = time.time()
        new_windows_handles = set()

        application_window_found = False
        main_app_pid = None

        while time.time() - start_time < max_wait_time:
            # Check for UAC prompt
            if handle_uac_prompt():
                logging.warning("Application triggered UAC prompt.")
                result['Status'] = 'Manual Intervention Required'
                result['Remarks'] = 'Application triggered UAC prompt.'
                result['Associated Windows'] = 'User Account Control'
                result['Terminated Executables'] = '; '.join(set(terminated_executables)) or 'None'
                result['Closed Windows'] = '; '.join(set(closed_windows)) or 'None'
                return result

            windows_after = set(w.handle for w in Desktop(backend="uia").windows())
            new_windows_handles = windows_after - windows_before

            # Try to find the expected window by executable name
            for handle in new_windows_handles:
                try:
                    window = Desktop(backend="uia").window(handle=handle)
                    process_id = window.process_id()
                    process = psutil.Process(process_id)
                    exe_path = process.exe()
                    if os.path.basename(exe_path).lower() == actual_executable_name.lower():
                        # Application window found
                        application_window_found = True
                        expected_window = window
                        main_app_pid = process_id
                        window_title = window.window_text()
                        associated_window_titles.append(window_title)
                        logging.info(f"Expected window detected: {window_title}")
                        # Pause after application window is found
                        pause_after_found = 2  # Adjust as needed
                        logging.info(f"Pausing for {pause_after_found} seconds to allow application to fully initialize.")
                        time.sleep(pause_after_found)
                        break  # Exit the wait loop
                except Exception as e:
                    logging.error(f"Error processing window handle {handle}: {e}")
                    continue

            if application_window_found:
                break  # Exit the wait loop

            # If not found by executable name, try matching by window title
            expected_title_pattern = re.escape(app_name)
            for handle in new_windows_handles:
                try:
                    window = Desktop(backend="uia").window(handle=handle)
                    window_title = window.window_text()
                    if re.search(expected_title_pattern, window_title, re.IGNORECASE):
                        # Application window found by title
                        application_window_found = True
                        expected_window = window
                        process_id = window.process_id()
                        main_app_pid = process_id
                        associated_window_titles.append(window_title)
                        logging.info(f"Expected window detected by title: {window_title}")
                        # Pause after application window is found
                        pause_after_found = 2  # Adjust as needed
                        logging.info(f"Pausing for {pause_after_found} seconds to allow application to fully initialize.")
                        time.sleep(pause_after_found)
                        break  # Exit the wait loop
                except Exception as e:
                    logging.error(f"Error processing window handle {handle}: {e}")
                    continue

            if application_window_found:
                break  # Exit the wait loop

            time.sleep(poll_interval)  # Wait before checking again

        if not application_window_found:
            # No application window opened within the wait time
            if new_windows_handles:
                # Handle new windows even if they don't match the expected executable
                for handle in new_windows_handles:
                    try:
                        window = Desktop(backend="uia").window(handle=handle)
                        window_title = window.window_text()
                        associated_window_titles.append(window_title)
                        window.close()
                        closed_windows.append(window_title)
                        logging.info(f"Closed window: {window_title}")
                    except Exception as e:
                        logging.error(f"Error processing window handle {handle}: {e}")
                        continue

                result['Status'] = 'Mostly Pass'
                result['Remarks'] = 'Expected application window not found, but other windows were handled.'
                result['Associated Windows'] = '; '.join(associated_window_titles)
                result['Terminated Executables'] = '; '.join(set(terminated_executables)) or 'None'
                result['Closed Windows'] = '; '.join(set(closed_windows)) or 'None'
                return result
            else:
                logging.error("No application windows opened after starting the application.")
                result['Status'] = 'Failed'
                result['Remarks'] = 'No application windows opened after starting the application.'
                result['Terminated Executables'] = '; '.join(set(terminated_executables)) or 'None'
                result['Closed Windows'] = '; '.join(set(closed_windows)) or 'None'
                return result

        # Wait an additional time to ensure the application has fully initialized
        additional_wait_time = 5  # Adjust as needed
        logging.info(f"Waiting an additional {additional_wait_time} seconds for application to fully initialize.")
        time.sleep(additional_wait_time)

        # Attempt to close the expected application window
        try:
            expected_window.close()
            closed_windows.append(expected_window.window_text())
            logging.info(f"Closed expected window: {expected_window.window_text()}")
        except Exception as e:
            logging.error(f"Error closing expected window: {e}")

        # Terminate the application process and its child processes
        if main_app_pid:
            logging.info(f"Terminating process tree starting with PID {main_app_pid}")
            terminated_execs = kill_process_tree(main_app_pid)
            terminated_executables.extend(terminated_execs)

        # Wait for processes to terminate
        time.sleep(2)

        # Verify the system has returned to its initial state
        processes_after = {p.pid for p in psutil.process_iter(['pid'])}
        new_pids = processes_after - processes_before

        for pid in new_pids:
            try:
                proc = psutil.Process(pid)
                proc_name = proc.name()
                if proc_name.lower() in ['svchost.exe', 'csrss.exe', 'wininit.exe', 'services.exe']:
                    logging.warning(f"Skipping termination of system process: PID {pid}, Name {proc_name}")
                    continue
                logging.warning(f"Process PID {pid} ({proc_name}) is still running after test.")
                # Attempt to terminate remaining processes
                proc.terminate()
                terminated_executables.append(proc_name)
            except psutil.NoSuchProcess:
                continue

        windows_after = set(w.handle for w in Desktop(backend="uia").windows())
        new_windows = windows_after - windows_before

        if new_windows:
            for handle in new_windows:
                try:
                    window = Desktop(backend="uia").window(handle=handle)
                    window_title = window.window_text()
                    logging.warning(f"Window '{window_title}' is still open after test.")
                    window.close()
                    closed_windows.append(window_title)
                except Exception as e:
                    logging.error(f"Error closing window handle {handle}: {e}")
                    continue

        # Add terminated executables and closed windows to result
        result['Terminated Executables'] = '; '.join(set(terminated_executables)) or 'None'
        result['Closed Windows'] = '; '.join(set(closed_windows)) or 'None'

        result['Status'] = 'Perfect Pass'
        result['Remarks'] = 'Expected application window opened and closed successfully.'
        result['Associated Windows'] = '; '.join(associated_window_titles)

    except Exception as e:
        result['Status'] = 'Failed'
        result['Remarks'] = f'Exception occurred: {e}'
        logging.error(f"Exception occurred during testing: {e}")
        result['Terminated Executables'] = '; '.join(set(terminated_executables)) or 'None'
        result['Closed Windows'] = '; '.join(set(closed_windows)) or 'None'

    return result

def save_results_to_excel(results, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Application Test Results"

    # Include new headers
    headers = [
        'Name', 'Shortcut Path', 'Expected Executable', 'Associated Windows',
        'Terminated Executables', 'Closed Windows', 'Status', 'Remarks'
    ]
    ws.append(headers)

    for result in results:
        ws.append([
            result.get('Name', ''),
            result.get('Shortcut Path', ''),
            result.get('Expected Executable', ''),
            result.get('Associated Windows', ''),
            result.get('Terminated Executables', ''),
            result.get('Closed Windows', ''),
            result.get('Status', ''),
            result.get('Remarks', '')
        ])

    # Adjust column widths for better readability
    for i, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in column:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set alignment for all cells to top-left
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

    wb.save(file_path)
    logging.info(f"Results saved to Excel file: {file_path}")

def main():
    global text_handler  # Declare as global to add/remove handler

    # Prompt the user to select the shortcuts and executables files
    root = tk.Tk()
    root.withdraw()

    # Directly open file dialogs without intermediate message boxes
    shortcuts_file = filedialog.askopenfilename(title="Select Shortcuts File", filetypes=(("Text files", "*.txt"),))
    if not shortcuts_file:
        logging.error("No shortcuts file selected.")
        print("No shortcuts file selected.")
        sys.exit(1)

    executables_file = filedialog.askopenfilename(title="Select Executables File", filetypes=(("Text files", "*.txt"),))
    if not executables_file:
        logging.error("No executables file selected.")
        print("No executables file selected.")
        sys.exit(1)

    # Load files before creating progress window
    shortcuts, executables = load_files(shortcuts_file, executables_file)
    if shortcuts is None or executables is None:
        sys.exit(1)

    # Create progress window
    total_apps = len(shortcuts)
    create_progress_window(total_apps)

    # Set up custom logging handler after GUI is initialized
    text_handler = TextWidgetHandler()
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    text_handler.setFormatter(formatter)
    logging.getLogger().addHandler(text_handler)

    results = []

    for index, (shortcut_path, exe_path) in enumerate(zip(shortcuts, executables), start=1):
        app_name = os.path.basename(shortcut_path).replace('.lnk', '')
        print(f"Testing application: {app_name}")
        logging.info(f"Testing application: {app_name}")

        update_progress_window(index, app_name)

        result = launch_and_test_application(shortcut_path, exe_path, app_name)
        results.append(result)
        time.sleep(2)

    # Close progress window
    progress_window.destroy()

    # Remove the custom logging handler
    logging.getLogger().removeHandler(text_handler)

    # Save results to Excel
    excel_file = os.path.join(os.getcwd(), 'Application_Test_Results.xlsx')
    save_results_to_excel(results, excel_file)
    print(f"Testing completed. Results saved to {excel_file}")
    logging.info("Testing completed.")

if __name__ == '__main__':
    main()
